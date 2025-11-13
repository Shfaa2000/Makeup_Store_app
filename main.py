from PyQt6 import QtCore, QtGui, QtWidgets
from PyQt6.QtWidgets import *
from PyQt6.QtGui import QFont, QIcon, QPixmap
from PyQt6.QtCore import Qt
import sys, os
from functools import partial
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from language_selector import LanguageDialog
from translations import translations


# دالة لتحديد مسار الموارد سواء في EXE أو ملف .py
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


class MakeupStore(QtWidgets.QMainWindow):
    def __init__(self, language):
        super().__init__()

        # -------- LANGUAGE ---------
        self.language = language
        self.tr = translations[self.language]

        # -------- WINDOW ---------
        self.setGeometry(10, 40, 1050, 680)
        self.setWindowTitle(self.tr["title"])
        self.setWindowIcon(QIcon(resource_path('images/logo.webp')))

        self.undo_stack = []
        self.inputs = []

        # -------- UI SETUP ---------
        self.setup_items_frame()
        self.setup_buying_buttons()
        self.setup_bill_frame()
        self.setup_main_frame()

    # ============================================================
    # --------------------- FRAME ITEMS --------------------------
    # ============================================================
    def setup_items_frame(self):
        self.frame_items = QFrame(self)
        self.frame_items.setGeometry(0, 0, 650, 680)
        self.frame_items.setStyleSheet('background-color: #f9f3f3')

        # Title
        title = QLabel(f"<b>{self.tr['title']}</b>", self.frame_items)
        title.setFixedSize(650, 50)
        title.setStyleSheet("""
        QLabel{
            font-size:22px;
            color:white;
            background-color: #d16b86;
            qproperty-alignment: 'AlignCenter';
            letter-spacing: 1px;
        }
        """)

        # Grid
        self.grid = QGridLayout(self.frame_items)
        self.grid.setContentsMargins(5, 10, 5, 5)
        self.grid.setHorizontalSpacing(4)
        self.grid.setVerticalSpacing(2)

        # Products
        self.products = [
            {"id": 1, "price": 40, "image": "images/baby makeup.jpg"},
            {"id": 2, "price": 20, "image": "images/eyeshadow.png"},
            {"id": 3, "price": 10, "image": "images/powder.png"},
            {"id": 4, "price": 50, "image": "images/foundation.jpg"},
            {"id": 5, "price": 30, "image": "images/highlighter.jpg"},
            {"id": 6, "price": 10, "image": "images/lip pencil.jpg"},
            {"id": 7, "price": 15, "image": "images/lipstick.jpg"},
            {"id": 8, "price": 60, "image": "images/makeupkit.jpg"},
            {"id": 9, "price": 35, "image": "images/mascara.jpg"},
            {"id": 10, "price": 25, "image": "images/lipglos.jpg"},
            {"id": 11, "price": 65, "image": "images/skincare.jpg"},
            {"id": 12, "price": 80, "image": "images/sunsceam.jpg"},
        ]

        BUTTON_WIDTH = 110
        BUTTON_HEIGHT = 120
        COLUMNS = 4

        for index, product in enumerate(self.products):
            row = index // COLUMNS
            col = index % COLUMNS

            button = QToolButton()
            button.setFixedSize(BUTTON_WIDTH, BUTTON_HEIGHT)
            button.setCursor(Qt.CursorShape.PointingHandCursor)

            icon = QIcon(resource_path(product["image"]))
            button.setIcon(icon)
            button.setIconSize(QtCore.QSize(90, 90))

            name = self.tr["products"][product["id"]]
            button.setText(name)
            button.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextUnderIcon)

            button.setStyleSheet("""
                QToolButton {
                    background-color: #f2d7d9;
                    border: none;
                    border-radius: 10px;
                    padding: 1px;
                    font-size: 12px;
                    color: #4e3b47;
                }
                QToolButton:hover {
                    background-color: #eac0c5;
                    border: 1px solid #d16b86;
                }
            """)

            button.clicked.connect(partial(self.add_to_bill, name, product["price"]))

            self.grid.addWidget(button, row, col)

    # ============================================================
    # ------------------- BUYING BUTTONS -------------------------
    # ============================================================
    def setup_buying_buttons(self):
        frame = QFrame(self)
        frame.setGeometry(0, 570, 650, 110)

        style = """
            QPushButton {
                background-color: #d16b86;
                font-size: 15px;
                color: white;
                border-radius: 8px;
            }
            QPushButton:hover {
                background-color: #b3546a;
            }
        """

        buttons = [
            (self.tr["buy_button"], self.buy_bill),
            (self.tr["new_bill"], self.new_bill),
            (self.tr["delete_item"], self.delete_selected_item),
            (self.tr["exit"], self.exit_program)
        ]

        x = 30
        for text, action in buttons:
            btn = QPushButton(text, frame)
            btn.setGeometry(x, 40, 140, 50)
            btn.setStyleSheet(style)
            btn.clicked.connect(action)
            x += 150

    # ============================================================
    # ---------------------- BILL FRAME --------------------------
    # ============================================================
    def setup_bill_frame(self):
        frame = QFrame(self)
        frame.setGeometry(650, 0, 400, 680)
        frame.setStyleSheet('background-color: white')

        layout = QVBoxLayout(frame)

        self.bill_tree = QTreeWidget()
        self.bill_tree.setHeaderLabels([
            self.tr["tree_items"],
            self.tr["tree_price"],
            self.tr["tree_quantity"],
            self.tr["tree_total"]
        ])
        self.bill_tree.setStyleSheet("font-size:12px;")
        self.bill_tree.header().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.bill_tree.header().setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.bill_tree)

    # ============================================================
    # ---------------------- MAIN FRAME --------------------------
    # ============================================================
    def setup_main_frame(self):
        frame = QFrame(self)
        frame.setGeometry(1050, 0, 300, 680)
        frame.setStyleSheet('background-color: #f2d7d9;')

        layout = QVBoxLayout(frame)
        labels = [
            self.tr["buyer_name"],
            self.tr["buyer_number"],
            self.tr["buyer_address"],
            self.tr["total_bill"],
            self.tr["buying_date"]
        ]

        for lab in labels:
            label = QLabel(lab)
            label.setStyleSheet("font-size:14px; color:#4e3b47;")
            input = QLineEdit()
            input.setStyleSheet("""
                background-color: white;
                padding: 6px;
                border: 1px solid #d1b6b6;
                border-radius: 5px;
            """)
            layout.addWidget(label)
            layout.addWidget(input)
            self.inputs.append(input)

        button_style = """
            QPushButton {
                padding: 10px;
                background-color: #d16b86;
                font-size: 14px;
                color: white;
                border-radius: 8px;
            }
            QPushButton:hover {
                background-color: #b3546a;
            }
        """

        btn_save = QPushButton(self.tr["save_bill"]); btn_save.setStyleSheet(button_style)
        btn_save.clicked.connect(self.save_bill)

        btn_empty = QPushButton(self.tr["empty_fields"]); btn_empty.setStyleSheet(button_style)
        btn_empty.clicked.connect(self.empty_fields)

        for btn in [btn_save, btn_empty]:
            layout.addWidget(btn)

        footer_image = QLabel()
        footer_image.setAlignment(Qt.AlignmentFlag.AlignCenter)
        pixmap = QPixmap(resource_path("images/salon.jpg"))
        pixmap = pixmap.scaledToWidth(270, Qt.TransformationMode.SmoothTransformation)
        footer_image.setPixmap(pixmap)
        footer_image.setStyleSheet("""
            border-radius: 8px;
            border: 2px solid #d16b86;
        """)
        layout.addWidget(footer_image)

    # ============================================================
    # ------------------------- LOGIC ----------------------------
    # ============================================================
    def add_to_bill(self, name, price):
        found_item = None
        for i in range(self.bill_tree.topLevelItemCount()):
            item = self.bill_tree.topLevelItem(i)
            if item.text(0) == name:
                found_item = item
                break

        if found_item:
            qty = int(found_item.text(2)) + 1
            found_item.setText(2, str(qty))
            found_item.setText(3, str(qty * price))
        else:
            item = QTreeWidgetItem([name, str(price), "1", str(price)])
            for col in range(4):
                item.setTextAlignment(col, Qt.AlignmentFlag.AlignCenter)
            self.bill_tree.addTopLevelItem(item)
            self.undo_stack.append(item)

    def buy_bill(self):
        total = 0
        for i in range(self.bill_tree.topLevelItemCount()):
            item = self.bill_tree.topLevelItem(i)
            total += float(item.text(3))

        self.inputs[3].setText(str(total) + " $")
        self.inputs[4].setText(datetime.now().strftime("%Y-%m-%d"))
        self.setGeometry(10, 40, 1350, 680)

    def new_bill(self):
        self.setGeometry(10, 40, 1050, 680)
        self.bill_tree.clear()

    def delete_selected_item(self):
        selected = self.bill_tree.selectedItems()
        if not selected:
            return
        item = selected[0]
        qty = int(item.text(2))
        if qty > 1:
            qty -= 1
            price = float(item.text(1))
            item.setText(2, str(qty))
            item.setText(3, str(qty * price))
        else:
            idx = self.bill_tree.indexOfTopLevelItem(item)
            if idx != -1:
                self.bill_tree.takeTopLevelItem(idx)

    def exit_program(self):
        QApplication.quit()

    def save_bill(self):
        name = self.inputs[0].text()
        phone = self.inputs[1].text()
        address = self.inputs[2].text()
        total = self.inputs[3].text()
        date = self.inputs[4].text()

        # التحقق من الحقول الفارغة
        for field in self.inputs:
            if field.text().strip() == "":
                QMessageBox.warning(self, self.tr["warning"], self.tr["sentence"])
                return

        file_path = "customers_information.xlsx"

        if not os.path.exists(file_path):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "customer"
            sheet["A1"] = "Full Name"
            sheet["B1"] = "Phone Number"
            sheet["C1"] = "Address"
            sheet["D1"] = "Total"
            sheet["E1"] = "Date"
            workbook.save(file_path)

        excel = openpyxl.load_workbook(file_path)
        sheet = excel.active
        row = sheet.max_row + 1

        sheet.cell(row=row, column=1, value=name)
        sheet.cell(row=row, column=2, value=phone)
        sheet.cell(row=row, column=3, value=address)
        sheet.cell(row=row, column=4, value=total)
        sheet.cell(row=row, column=5, value=date)

        excel.save(file_path)
        QMessageBox.information(self, self.tr["success"], self.tr["save"])

    def empty_fields(self):
        for field in self.inputs:
            field.clear()


# =====================================================================
# --------------------------- PROGRAM START ---------------------------
# =====================================================================

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setFont(QFont('Poppins'))

    # language selector
    lang_dialog = LanguageDialog()
    if lang_dialog.exec() == LanguageDialog.DialogCode.Accepted:
        language = lang_dialog.selected_language
        if language == "ar":
            app.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        else:
            app.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
    else:
        sys.exit()

    window = MakeupStore(language)
    window.show()
    app.exec()
